import pandas as pd
import math
import difflib
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ==========================================
# CONFIGURATION DES FICHIERS LOCAUX
# ==========================================
FICHIER_ERP = "Book2.xlsx"
FICHIER_EXTRACTION_AZURE = "Bilan_et_Extraction_SDC_V8.xlsx" # À adapter
OUTPUT_EXCEL = "Nouveau_Bilan_Test_V18.xlsx"

# Couleurs pour le formatage Excel
FILL_VERT = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_ORANGE = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_ROUGE = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

def safe_float(val):
    if pd.isna(val) or val == "": return 0.0
    try:
        if isinstance(val, str):
            val = val.replace(',', '.').replace('€', '').replace('EUR', '').replace(' ', '').strip()
        return float(val)
    except:
        return 0.0

def safe_float_erp_price(val):
    num = safe_float(val)
    if num >= 10000: num = num / 1000000.0
    return num

def load_and_group_erp(filepath):
    df = pd.read_excel(filepath)
    df['No. de Cde Interne'] = df['No. de Cde Interne'].astype(str).str.replace(" ", "").str.upper()
    df['Article - Référence'] = df['Article - Référence'].fillna('')
    df['Article - Nom'] = df['Article - Nom'].fillna('')
    
    df['Quantité'] = df['Quantité'].apply(safe_float)
    df['P.U. - Cde Fournisseur'] = df['P.U. - Cde Fournisseur'].apply(safe_float_erp_price)
    
    grouped = df.groupby(
        ['No. de Cde Interne', 'Article - Référence', 'Article - Nom', 'P.U. - Cde Fournisseur'], 
        dropna=False
    )['Quantité'].sum().reset_index()
    return grouped

def match_score(erp_ref, erp_nom, erp_pu, pdf_desc, pdf_pu):
    score = 0
    pdf_desc_str = str(pdf_desc).lower()
    if erp_ref and str(erp_ref).lower() in pdf_desc_str: return 100
    if erp_pu > 0 and math.isclose(erp_pu, pdf_pu, abs_tol=0.01): score += 80
    if erp_nom:
        text_sim = difflib.SequenceMatcher(None, str(erp_nom).lower(), pdf_desc_str).ratio()
        score += (text_sim * 50)
    return score

def main():
    print("🚀 Démarrage du simulateur de rapprochement V18 (Tolérance 20% & Tri Numérique)")
    
    try:
        df_erp = load_and_group_erp(FICHIER_ERP)
        print("✅ Base SDC chargée.")
    except Exception as e:
        print(f"❌ Erreur lecture SDC : {e}"); return

    try:
        df_brut = pd.read_excel(FICHIER_EXTRACTION_AZURE, sheet_name="Extraction_Brute_Azure")
        df_brut = df_brut.fillna("")
        print("✅ Données brutes Azure chargées.")
    except Exception as e:
        print(f"❌ Impossible de lire l'onglet Azure : {e}"); return

    bilan_data = []
    groupes_factures = df_brut.groupby('Fichier')

    for file_name, group in groupes_factures:
        first_row = group.iloc[0]
        refs_str = str(first_row['Refs_Interne_Lues'])
        refs_internes_list = [r.strip() for r in refs_str.split(",")] if (refs_str and refs_str != "Non Trouvé") else []
        ref_externe = first_row['Num_Facture']
        date_facture = first_row['Date_Facture']
        total_pdf_brut = safe_float(first_row['Montant_Total_Lu'])

        # RÈGLE 1 : Aucun CLX
        if len(refs_internes_list) == 0:
            bilan_data.append({
                "Fichier_PDF": file_name, "Reference": "NON_TROUVE", "Ligne_Num": "", "Article": "AUCUN_CLX_DETECTE", 
                "Quantite": 0, "Quantite_Lue": 0, "Prix_Unitaire": 0, "Prix_Unitaire_Lu": 0, 
                "Prix_Transport_Unitaire": 0, "Ref_Facture_Externe": ref_externe, 
                "Date_Facture": date_facture, "Statut_Validation": "6_ERREUR_AUCUN_CLX", 
                "Validation_Globale": "NON", "Alerte_Transport": "N/A"
            })
            continue

        # RÈGLE 2 : Multi-CLX
        if len(refs_internes_list) > 1:
            bilan_data.append({
                "Fichier_PDF": file_name, "Reference": ", ".join(refs_internes_list), "Ligne_Num": "", "Article": "FACTURE_MULTI_COMMANDES", 
                "Quantite": 0, "Quantite_Lue": 0, "Prix_Unitaire": 0, "Prix_Unitaire_Lu": 0, 
                "Prix_Transport_Unitaire": 0, "Ref_Facture_Externe": ref_externe, 
                "Date_Facture": date_facture, "Statut_Validation": "6_ERREUR_MULTI_CLX", 
                "Validation_Globale": "NON", "Alerte_Transport": "N/A"
            })
            continue

        # RÈGLE 3 : Traitement normal 1 CLX
        ref_commande = refs_internes_list[0]
        lignes_erp = df_erp[df_erp['No. de Cde Interne'] == ref_commande]
        
        if lignes_erp.empty:
            bilan_data.append({
                "Fichier_PDF": file_name, "Reference": ref_commande, "Ligne_Num": "", "Article": "INCONNU_DANS_SDC", 
                "Quantite": 0, "Quantite_Lue": 0, "Prix_Unitaire": 0, "Prix_Unitaire_Lu": 0, 
                "Prix_Transport_Unitaire": 0, "Ref_Facture_Externe": ref_externe, 
                "Date_Facture": date_facture, "Statut_Validation": "6_ERREUR_CLX_INCONNU", 
                "Validation_Globale": "NON", "Alerte_Transport": "N/A"
            })
            continue

        lignes_pdf_brutes = []
        for _, row in group.iterrows():
            if str(row['Ligne_Num']).strip():
                qte = safe_float(row['Qte_Lue'])
                pu_azure = safe_float(row['PU_Lu'])
                montant = safe_float(row['Montant_Ligne_Lu'])
                pu_calcule = pu_azure
                
                if qte > 0 and montant > 0:
                    pu_theorique = montant / qte
                    if pu_azure == 0 or pu_azure > (pu_theorique * 5) or pu_azure < (pu_theorique / 5):
                        pu_calcule = pu_theorique
                lignes_pdf_brutes.append({"desc": str(row['Description_Lue']), "qte": qte, "pu": pu_calcule, "montant": montant})

        pdf_fusion_dict = {}
        for l in lignes_pdf_brutes:
            cle = (l['desc'].strip().lower(), round(l['pu'], 2))
            if cle not in pdf_fusion_dict:
                pdf_fusion_dict[cle] = l.copy()
                pdf_fusion_dict[cle]["matched"] = False
            else:
                pdf_fusion_dict[cle]['qte'] += l['qte']
                pdf_fusion_dict[cle]['montant'] += l['montant']
        lignes_pdf = list(pdf_fusion_dict.values())

        qte_totale_globale = 0.0
        total_attendu_global = 0.0
        match_results = []
        
        for index, row_erp in lignes_erp.iterrows():
            nom_erp, ref_erp = str(row_erp['Article - Nom']), str(row_erp['Article - Référence'])
            pu_erp, qte_erp = safe_float_erp_price(row_erp['P.U. - Cde Fournisseur']), safe_float(row_erp['Quantité'])
            
            qte_totale_globale += qte_erp
            total_attendu_global += (pu_erp * qte_erp)
            
            pdf_qte, pdf_pu, meilleur_match_idx = 0, 0, -1
            
            if len(lignes_erp) == 1 and lignes_pdf:
                for i, l_pdf in enumerate(lignes_pdf):
                    if math.isclose(pu_erp, l_pdf["pu"], abs_tol=0.01): meilleur_match_idx = i; break
                if meilleur_match_idx == -1: meilleur_match_idx = max(range(len(lignes_pdf)), key=lambda i: lignes_pdf[i]["montant"])
            else:
                meilleur_score = 0
                for i, l_pdf in enumerate(lignes_pdf):
                    if not l_pdf["matched"]:
                        score = match_score(ref_erp, nom_erp, pu_erp, l_pdf["desc"], l_pdf["pu"])
                        if score > 30 and score > meilleur_score: meilleur_score, meilleur_match_idx = score, i
            
            if meilleur_match_idx != -1:
                lignes_pdf[meilleur_match_idx]["matched"] = True
                pdf_qte, pdf_pu = lignes_pdf[meilleur_match_idx]["qte"], lignes_pdf[meilleur_match_idx]["pu"]
            
            match_results.append({
                "Fichier_PDF": file_name, "Reference_Cde": ref_commande, "Article_ERP": f"{ref_erp} - {nom_erp}",
                "Quantite_ERP": qte_erp, "Quantite_PDF": pdf_qte, "PU_ERP": pu_erp, "PU_PDF": pdf_pu,
                "Ref_Facture_Externe": ref_externe, "Date_Facture": date_facture
            })

        # Calcul du transport et de l'Alerte
        montant_lignes_matchees = sum(l["montant"] for l in lignes_pdf if l["matched"])
        surcharge_globale = total_pdf_brut - montant_lignes_matchees
        if surcharge_globale < 0.50: surcharge_globale = 0.0
        transport_unitaire = surcharge_globale / qte_totale_globale if qte_totale_globale > 0 else 0

        # Alerte si les frais de port représentent plus de 20% de la valeur de la commande
        alerte_transport = "OK"
        if surcharge_globale > (0.20 * total_attendu_global) and surcharge_globale > 5.0:
            alerte_transport = "ANORMAL (>20%)"

        for i, res in enumerate(match_results, 1):
            qte_ok = math.isclose(res["Quantite_ERP"], res["Quantite_PDF"], abs_tol=0.01)
            pu_ok = math.isclose(res["PU_ERP"], res["PU_PDF"], abs_tol=0.01)
            
            ecart_pu_pct = 0
            if res["PU_ERP"] > 0:
                ecart_pu_pct = abs(res["PU_PDF"] - res["PU_ERP"]) / res["PU_ERP"]

            # Logique de Classification Stricte
            if res["Quantite_PDF"] == 0: 
                statut = "5_NON_TROUVE"
            elif not qte_ok: 
                statut = "4_ECART_QUANTITE"
            elif pu_ok: 
                statut = "1_PARFAIT"
            elif ecart_pu_pct <= 0.20: 
                statut = "2_ECART_PRIX_MINEUR"
            else: 
                statut = "3_ECART_PRIX_MAJEUR"

            # Validation automatique pour les parfaits et les écarts mineurs
            est_valide = "OUI" if statut in ["1_PARFAIT", "2_ECART_PRIX_MINEUR"] else "NON"

            bilan_data.append({
                "Fichier_PDF": res["Fichier_PDF"], "Reference": res["Reference_Cde"], "Ligne_Num": i,
                "Article": res["Article_ERP"], "Quantite": res["Quantite_ERP"], "Quantite_Lue": res["Quantite_PDF"],
                "Prix_Unitaire": res["PU_ERP"], "Prix_Unitaire_Lu": res["PU_PDF"],
                "Prix_Transport_Unitaire": round(transport_unitaire, 4),
                "Ref_Facture_Externe": res["Ref_Facture_Externe"], "Date_Facture": res["Date_Facture"],
                "Statut_Validation": statut, "Validation_Globale": est_valide, "Alerte_Transport": alerte_transport
            })

    if not bilan_data: return

    df_bilan = pd.DataFrame(bilan_data)
    df_bilan.to_excel(OUTPUT_EXCEL, sheet_name="Bilan_Test", index=False)

    wb = load_workbook(OUTPUT_EXCEL)
    ws_bilan = wb["Bilan_Test"]

    for row_idx, data in enumerate(bilan_data, start=2):
        cell_qte, cell_qte_lue = ws_bilan.cell(row=row_idx, column=5), ws_bilan.cell(row=row_idx, column=6)
        cell_pu, cell_pu_lu = ws_bilan.cell(row=row_idx, column=7), ws_bilan.cell(row=row_idx, column=8)
        cell_statut = ws_bilan.cell(row=row_idx, column=12) 
        cell_val_glob = ws_bilan.cell(row=row_idx, column=13) 
        cell_alerte_trans = ws_bilan.cell(row=row_idx, column=14)
        
        statut = data.get("Statut_Validation", "")

        # Mise en forme : Statut & Validation
        if statut == "1_PARFAIT": 
            cell_statut.fill = cell_val_glob.fill = FILL_VERT
        elif statut == "2_ECART_PRIX_MINEUR":
            cell_statut.fill = FILL_ORANGE # On signale que c'est un écart
            cell_val_glob.fill = FILL_VERT   # MAIS il est validé automatiquement
        elif statut in ["3_ECART_PRIX_MAJEUR", "4_ECART_QUANTITE"]: 
            cell_statut.fill = cell_val_glob.fill = FILL_ORANGE
        else: 
            cell_statut.fill = cell_val_glob.fill = FILL_ROUGE 

        # Mise en forme : Alerte Transport
        if data.get("Alerte_Transport") == "ANORMAL (>20%)":
            cell_alerte_trans.fill = FILL_ROUGE
        elif data.get("Alerte_Transport") == "OK":
            cell_alerte_trans.fill = FILL_VERT

        # Mise en forme : Cellules QTE et PU
        if statut in ["1_PARFAIT", "2_ECART_PRIX_MINEUR", "3_ECART_PRIX_MAJEUR", "4_ECART_QUANTITE", "5_NON_TROUVE"]:
            if math.isclose(data["Quantite"], data["Quantite_Lue"], abs_tol=0.01): 
                cell_qte.fill = cell_qte_lue.fill = FILL_VERT
            elif data["Quantite_Lue"] == 0: 
                cell_qte.fill = cell_qte_lue.fill = FILL_ROUGE
            else: 
                cell_qte.fill = cell_qte_lue.fill = FILL_ORANGE
                
            if math.isclose(data["Prix_Unitaire"], data["Prix_Unitaire_Lu"], abs_tol=0.01): 
                cell_pu.fill = cell_pu_lu.fill = FILL_VERT
            elif data["Prix_Unitaire_Lu"] == 0: 
                cell_pu.fill = cell_pu_lu.fill = FILL_ROUGE
            else: 
                cell_pu.fill = cell_pu_lu.fill = FILL_ORANGE

    wb.save(OUTPUT_EXCEL)
    print(f"✅ Bilan V18 généré : {OUTPUT_EXCEL}")

if __name__ == "__main__":
    main()